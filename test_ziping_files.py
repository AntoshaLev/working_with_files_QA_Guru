import os
from utils import RESOURCES_PATH, TMP_PATH
from zipfile import ZipFile
from xlrd import open_workbook
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader


def test_zip_from_resources_names():
    """Проверяем архив на наличие всех файлов по списку имен"""
    # получаем список имен файлов в resources
    file_in_dir = os.listdir(RESOURCES_PATH)
    with ZipFile(os.path.join(TMP_PATH, 'zip_arh_in_tests.zip')) as zf:
        assert file_in_dir == zf.namelist()


@pytest.mark.parametrize("file_name", ['file_or_text.txt'])
def test_txt_in_zip(file_name):
    # получаем данные файла и его содержимое
    txt_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    with open(os.path.join(RESOURCES_PATH, file_name)) as f:
        txt_file_text = f.read()

    with ZipFile(os.path.join(TMP_PATH, 'zip_arh_in_tests.zip')) as zf:
        # сравниваем данные файла в архиве с полученными
        assert txt_file_size == zf.getinfo(file_name).file_size
        assert txt_file_text in zf.read(file_name).decode('utf-8')


@pytest.mark.parametrize("file_name", ['file_example_XLS_10.xls'])
def test_xls_in_zip(file_name):
    # получаем данные файла и его содержимое
    xls_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    book = open_workbook(os.path.join(RESOURCES_PATH, file_name))
    sheets_count = book.nsheets
    sheets_names = book.sheet_names()
    xls_file_text = book.sheet_by_index(0).cell_value(8, 1)

    with ZipFile(os.path.join(TMP_PATH, 'zip_arh_in_tests.zip')) as zf:
        # сравниваем данные файла в архиве с полученными
        assert xls_file_size == zf.getinfo(file_name).file_size
        book_zip = open_workbook(file_contents=zf.read(file_name))
        assert sheets_count == book_zip.nsheets
        assert sheets_names == book_zip.sheet_names()
        assert xls_file_text == book_zip.sheet_by_index(0).cell_value(8, 1)


@pytest.mark.parametrize("file_name", ['file_example_XLSX_50.xlsx'])
def test_xlsx_in_zip(file_name):
    # получаем данные файла и его содержимое
    xlsx_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    book = load_workbook(os.path.join(RESOURCES_PATH, file_name))
    sheets_count = len(book.sheetnames)
    sheets_names = book.sheetnames
    sheet = book.active
    xlsx_file_text = sheet.cell(9, 3).value

    with ZipFile(os.path.join(TMP_PATH, 'zip_arh_in_tests.zip')) as zf:
        # сравниваем данные файла в архиве с полученными
        assert xlsx_file_size == zf.getinfo(file_name).file_size
        book_zip = load_workbook(zf.open(file_name, 'r'))
        assert sheets_count == len(book_zip.sheetnames)
        assert sheets_names == book_zip.sheetnames
        sheet_zip = book_zip.active
        assert xlsx_file_text == sheet_zip.cell(9, 3).value


@pytest.mark.parametrize("file_name", ['Python Testing with Pytest (Brian Okken).pdf'])
def test_pdf_in_zip(file_name):
    # получаем данные файла и его содержимое
    pdf_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))

    reader = PdfReader(os.path.join(RESOURCES_PATH, file_name))
    pdf_page_count = len(reader.pages)
    pdf_page_text = reader.pages[1].extract_text()

    with ZipFile(os.path.join(TMP_PATH, 'zip_arh_in_tests.zip'), mode='r') as zf:
        # сравниваем данные файла в архиве с полученными
        assert pdf_file_size == zf.getinfo(file_name).file_size
        zip_reader = PdfReader(zf.open(file_name, 'r'))
        assert pdf_page_count == len(zip_reader.pages)
        assert pdf_page_text == zip_reader.pages[1].extract_text()
